# -*- coding: utf-8 -*-
"""Scan from customer/SAP data for tab-export .xls and real .xlsx; emit markdown."""
from __future__ import annotations

import json
import re
from collections import defaultdict
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parents[1]
ROOT = REPO / "from customer" / "SAP data"
DOCS = REPO / "docs"

KEYWORDS = (
    "Order",
    "Material",
    "Equipment",
    "Functional",
    "Plant",
    "Posting",
    "Work",
    "OrdCat",
    "Descr",
    "Type",
    "Status",
    "Movement",
    "Qty",
    "Batch",
    "Storage",
    "Cost",
    "Ctr",
    "Confirm",
    "Issue",
    "Receipt",
    "GI",
    "GR",
)


def decode_text(path: Path) -> str:
    raw = path.read_bytes()
    if raw.startswith(b"\xff\xfe"):
        return raw.decode("utf-16-le")
    if raw.startswith(b"\xfe\xff"):
        return raw.decode("utf-16-be")
    for enc in ("utf-8-sig", "utf-8", "cp874", "latin-1"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def _mostly_numeric(parts: list[str]) -> bool:
    if not parts:
        return True
    num = 0
    for p in parts:
        s = p.replace(",", "").strip()
        if re.fullmatch(r"[\d.Ee+-]+", s) or (s.isdigit() and len(s) >= 6):
            num += 1
    return num / len(parts) > 0.35


def best_tab_header_line(text: str) -> list[str]:
    best: list[str] = []
    best_score = 0
    for line in text.splitlines():
        if "\t" not in line:
            continue
        parts = [p.strip() for p in line.split("\t")]
        parts = [p for p in parts if p]
        if len(parts) < 5:
            continue
        if len(parts) <= 2 and max(len(p) for p in parts) > 120:
            continue
        if _mostly_numeric(parts):
            continue
        kw = sum(1 for p in parts for k in KEYWORDS if k.lower() in p.lower())
        numish = sum(1 for p in parts if re.fullmatch(r"[\d.,:\sEe+-]+", p))
        score = len(parts) + kw * 3 - numish
        if score > best_score:
            best_score = score
            best = parts
    return best


def headers_xlsx(path: Path) -> list[dict]:
    out: list[dict] = []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for name in wb.sheetnames:
            ws = wb[name]
            hdrs: list[str] = []
            for r in range(1, 8):
                row = next(ws.iter_rows(min_row=r, max_row=r, values_only=True), None)
                if not row:
                    continue
                hdrs = [str(c).strip() for c in row if c is not None and str(c).strip()]
                if len(hdrs) >= 3:
                    break
            out.append({"sheet": name, "headers": hdrs[:120]})
        wb.close()
    except Exception as e:
        out.append({"sheet": "_error_", "headers": [str(e)]})
    return out


def headers_file(path: Path) -> list[dict]:
    suf = path.suffix.lower()
    if suf == ".xlsx":
        return headers_xlsx(path)
    if suf == ".xls":
        text = decode_text(path)
        hdr = best_tab_header_line(text)
        return [{"sheet": "(SAP list export, tab-separated)", "headers": hdr}]
    return []


def classify(path: str) -> str:
    p = path.replace("\\", "/").lower()
    base = Path(path).name.lower()
    if "iw37n" in p:
        return "IW37N"
    if base.startswith("gi ") or "/gi " in p or p.endswith("/gi.xls"):
        return "GI"
    if base.startswith("gr ") or "/gr " in p or p.endswith("/gr.xls"):
        return "GR"
    if "confirm" in p:
        return "Confirm WO"
    if "functional" in p or "equipment" in p:
        return "Functional location & equipment"
    if "ip19" in p:
        return "IP19"
    if "ia17" in p:
        return "IA17"
    if "work center" in p:
        return "Work center list"
    if "task list" in p or "gen task" in p or "general task" in p:
        return "Task list"
    if "pc50" in p:
        return "PC50"
    if "pm database" in p or "proposal" in p or "status wo" in p:
        return "PM / WO supporting"
    return "Other"


def main() -> None:
    results: list[dict] = []
    for p in sorted(ROOT.rglob("*")):
        if not p.is_file():
            continue
        if p.suffix.lower() not in (".xls", ".xlsx"):
            continue
        sheets = headers_file(p)
        try:
            rel_path = p.relative_to(REPO).as_posix()
        except ValueError:
            rel_path = p.as_posix()
        results.append(
            {
                "path": rel_path,
                "category": classify(p.as_posix()),
                "sheets": sheets,
            }
        )

    DOCS.mkdir(parents=True, exist_ok=True)
    dbg = Path(__file__).resolve().parent / "_sap_columns_raw.json"
    dbg.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")

    by_cat: dict[str, list[tuple[str, list[str]]]] = defaultdict(list)
    for r in results:
        hdr = r["sheets"][0]["headers"] if r["sheets"] else []
        key = r["category"]
        by_cat[key].append((r["path"], hdr))

    lines: list[str] = []
    lines.append("# SAP import / export — คอลัมน์จากตัวอย่างใน `from customer/SAP data`")
    lines.append("")
    lines.append(
        "ไฟล์ `.xls` หลายไฟล์เป็น **SAP ALV / list export แบบ tab-separated** (บันทึกนามสกุล `.xls`) ไม่ใช่ BIFF — สคริปต์ดึงแถวหัวตารางโดยประมาณจากแถวที่มีคีย์เวิร์ดคอลัมน์มากที่สุด"
    )
    lines.append(
        "รูปแบบคอลัมน์ **อาจต่างกันระหว่างไฟล์/รุ่นรายงาน SAP** (เช่น IW37N หลายชุด) — ใช้ตารางด้านล่างเป็น baseline แล้วตรวจสอบกับไฟล์จริงที่นำเข้าในแต่ละ batch"
    )
    lines.append("")
    lines.append("- **นำเข้า (import จาก SAP / ไฟล์ที่ SAP ส่งออก):** คอลัมน์ด้านล่างคือสิ่งที่ปรากฏในตัวอย่าง — ใช้เป็น data dictionary ฝั่งแอป/staging")
    lines.append("- **ส่งออกกลับ SAP (export):** โรงงานมักใช้ BAPI / LSMW / template ที่กำหนดใน SAP — ไฟล์ในโฟลเดอร์นี้ส่วนใหญ่เป็น **ตัวอย่างผลลัพธ์จาก SAP** มากกว่าเทมเพลตนำเข้า; การจับคู่กลับ SAP ต้องยืนยันกับทีม MM/PM อีกครั้ง")
    lines.append("")
    lines.append("สร้าง/อัปเดตอัตโนมัติได้ด้วย `python scripts/extract_sap_data_columns.py`")
    lines.append("")
    lines.append("อ้างอิงเพิ่ม: [`DATABASE_DESIGN_DRAFT.md`](DATABASE_DESIGN_DRAFT.md), [`CUSTOMER_FROM_FOLDER_MANIFEST.md`](CUSTOMER_FROM_FOLDER_MANIFEST.md)")
    lines.append("")

    for cat in sorted(by_cat.keys(), key=lambda x: (x == "Other", x)):
        lines.append(f"## {cat}")
        lines.append("")
        if cat == "Functional location & equipment":
            lines.append(
                "ไฟล์ **Functional Location & Equipment.xls** เป็น **โครงสร้างลำดับชั้น** (location / equipment / คำอธิบาย) มากกว่าตารางแบน — ถ้าไม่มีแถวหัวที่อ่านได้ ให้ออกแบบ import แบบ tree หรือ flatten (เช่น `floc_id`, `parent_id`, `equipment_id`, `description`)"
            )
            lines.append("")
        seen_hdr: set[tuple[str, ...]] = set()
        for path, hdr in by_cat[cat]:
            if not hdr or (len(hdr) == 1 and hdr[0].startswith("_")):
                lines.append(f"- `{path}` — *ไม่พบแถวหัวที่อ่านได้*")
                lines.append("")
                continue
            t = tuple(hdr)
            if t in seen_hdr:
                lines.append(f"- `{path}` — *หัวตารางซ้ำกับไฟล์อื่นในกลุ่มนี้*")
                lines.append("")
                continue
            seen_hdr.add(t)
            lines.append(f"### `{path}`")
            lines.append("")
            lines.append("| # | Column (as in file) |")
            lines.append("|:-:|----------------------|")
            for i, col in enumerate(hdr, 1):
                safe = col.replace("|", "\\|")
                lines.append(f"| {i} | {safe} |")
            lines.append("")

    out_path = DOCS / "SAP_DATA_IMPORT_EXPORT_COLUMNS.md"
    out_path.write_text("\n".join(lines), encoding="utf-8")
    print("Wrote", out_path, "categories", len(by_cat))


if __name__ == "__main__":
    main()
